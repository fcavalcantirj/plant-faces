"""Thales.xlsx -> load-ready NDJSON backfill (T0.11).

Months of real grow air data (Tuya T/H sensor -> Google Apps Script -> Sheets)
live in monthly tabs. This one-shot importer turns them into archive envelope
rows: source `tuya-sheets-backfill`, channels `air_temp_c`/`rh_pct`, flagged
uncalibrated and `ingest_source: backfill`. The archive DB is not live yet
(Felipe gate), so the actual load is DEFERRED — the NDJSON written here is the
load-ready artifact, one envelope per line.

Columns are taken POSITIONALLY (Data, Temp, Umid, VPD, Alarme*): header rows
carry extra running-tally cells to the right ("VPD Correto (OK)" counts) which
are ignored. The Alarme header names the stage band in force ("Alarme Vega" /
"Alarme Flora") — the cell text is carried verbatim as `meta.alarm_raw`.

Timestamps: the Data column is an Excel serial date; openpyxl converts it with
the workbook's 1899-12-30 epoch (asserted — a date1904 workbook would silently
shift everything by 4 years). The wall clock is America/Sao_Paulo, which is
UTC-3 year-round (Brazil abolished DST in 2019; the sheets span 2026-06..08,
all UTC-3), so the conversion is a fixed +3h: wall clock -> true UTC epoch ms.
`meta.tz_note` records the conversion applied on every row.

Rows where Temp is error text ("Erro", "Erro na API", ... — the Tuya
cloud-plan/token failures visible in Tempo Real) are SKIPPED and counted, per
error string. Fully blank rows are counted. Any OTHER shape (partial row,
non-datetime Data) is refused loudly — an unclassifiable row is a bug in our
understanding of the sheet, not a row to guess about.

Validation: VPD is recomputed (etl/vpd.py, Tetens) for every valid row and
diffed against the sheet's own VPD column. Two sheet-VPD encodings exist:
plain "9.99" text, and cells the Sheets->xlsx export locale-mangled into
DATES — "1.06" became 2026-06-01 (day.month) — decoded back as day + month/100
only for midnight-2026 cells; anything else is refused, never guessed.

    .venv/bin/python -m etl.backfill_thales [Thales.xlsx] [out.ndjson]
"""

import calendar
import datetime
import json
import math
import sys
from pathlib import Path

import openpyxl

from etl.vpd import vpd_air_kpa

SOURCE_TOKEN = "tuya-sheets-backfill"

# Workbook tab order is 07, 08, 06, Tempo Real; we process chronologically
# (Tempo Real 05-01..06-21, then the monthly tabs). Verified 2026-08-08: the
# ranges are disjoint — no timestamp appears in two sheets — and a global
# uniqueness guard below keeps that true against future re-exports.
SHEETS = ("Tempo Real", "062026", "072026", "082026")

# Positional header prefix every sheet must present; the 5th column is the
# stage-band alarm header ("Alarme Vega" / "Alarme Flora").
HEADER_PREFIX = ("Data", "Temp", "Umid")
EXCEL_EPOCH_1900 = datetime.datetime(1899, 12, 30)

# Sheet VPD is rounded to 2dp, so a same-formula recompute lands within
# +/-0.005; the wider gate absorbs float artifacts without hiding a different
# formula, which would sit far outside it.
VPD_TOLERANCE_KPA = 0.02

# America/Sao_Paulo is UTC-3 year-round (Brazil abolished DST in 2019); the
# sheets span 2026-06..08, so a single fixed offset covers every row.
SP_UTC_OFFSET_MS = 3 * 60 * 60 * 1000  # 10_800_000

TZ_NOTE = "wall clock America/Sao_Paulo (UTC-3, no DST) converted to UTC epoch ms"

# Histogram bin edges for (recomputed - sheet) diffs, kPa.
HIST_EDGES = (-0.02, -0.01, -0.005, 0.005, 0.01, 0.02)


def to_epoch_ms_utc(dt: datetime.datetime) -> int:
    """Naive America/Sao_Paulo wall-clock datetime -> true UTC epoch ms.

    Sao Paulo is UTC-3 year-round (no DST since 2019), so wall clock -> UTC is
    a fixed +3h. Integer math (timegm), not timestamp(): the host's local zone
    must never leak into the encoding. A tz-aware datetime is refused — this
    sheet has no offset on record, so one showing up means the parse changed
    under us.
    """
    if dt.tzinfo is not None:
        raise ValueError(f"expected naive wall-clock datetime, got tz-aware {dt!r}")
    return calendar.timegm(dt.timetuple()) * 1000 + dt.microsecond // 1000 + SP_UTC_OFFSET_MS


def decode_sheet_vpd(cell) -> float | None:
    """Sheet VPD cell -> kPa float, or None where the sheet has no value.

    Plain "9.99" text parses directly. The date-mangled cells (Sheets->xlsx
    export read "1.06" as day-1 month-06) decode as day + month/100, but ONLY
    for midnight cells in the mangle year 2026 — any other datetime shape is
    not a known mangle and is refused rather than decoded on a hunch.
    """
    if cell is None:
        return None
    if isinstance(cell, str):
        try:
            return float(cell)
        except ValueError:
            raise ValueError(f"unparseable sheet VPD text {cell!r}") from None
    if isinstance(cell, datetime.datetime):
        if cell.year == 2026 and (cell.hour, cell.minute, cell.second, cell.microsecond) == (0, 0, 0, 0):
            return cell.day + cell.month / 100.0
        raise ValueError(f"sheet VPD date-cell {cell!r} outside the known day.month mangle")
    if isinstance(cell, (int, float)):
        return float(cell)
    raise ValueError(f"sheet VPD cell of unexpected type {type(cell).__name__}: {cell!r}")


def classify_row(data, temp, umid) -> str:
    """'valid' | 'error' | 'blank' — anything else raises.

    Error rows carry the failure text in Temp (Umid holds the detail); both
    are strings together. A row that fits none of the three shapes is refused:
    guessing here is how bad rows become dataset rows.
    """
    if data is None and temp is None and umid is None:
        return "blank"
    if isinstance(data, datetime.datetime) and isinstance(temp, str) and isinstance(umid, str):
        return "error"
    if isinstance(data, datetime.datetime) and isinstance(temp, (int, float)) and isinstance(umid, (int, float)):
        return "valid"
    raise ValueError(f"unclassifiable row shape: data={data!r} temp={temp!r} umid={umid!r}")


def envelope(ts_ms: int, temp_c: float, rh_pct: float, sheet: str, alarm_raw) -> dict:
    return {
        "source_token": SOURCE_TOKEN,
        "ts": ts_ms,
        "channels": {"air_temp_c": temp_c, "rh_pct": rh_pct},
        "meta": {
            "adapter": "sheets-backfill",
            "calibration": "uncalibrated",
            "ingest_source": "backfill",
            "sheet": sheet,
            "alarm_raw": alarm_raw,
            "tz_note": TZ_NOTE,
        },
    }


def parse_sheet(ws, sheet_name: str) -> tuple[list[dict], dict]:
    """One worksheet -> (envelope rows, per-sheet report dict)."""
    stats = {
        "sheet": sheet_name,
        "total_data_rows": 0,
        "valid": 0,
        "blank": 0,
        "errors": {},
        "with_sheet_vpd": 0,
        "vpd_match": 0,
        "hist": [0] * (len(HIST_EDGES) + 1),
        "beyond_tolerance": [],
    }
    rows: list[dict] = []
    for i, raw in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = (list(raw) + [None] * 5)[:5]
        data, temp, umid, vpd_cell, alarm = cells
        if i == 1:
            if tuple(cells[:3]) != HEADER_PREFIX or not (isinstance(alarm, str) and alarm.startswith("Alarme")):
                raise ValueError(f"{sheet_name}: unexpected header {cells!r}")
            continue

        kind = classify_row(data, temp, umid)
        if kind == "blank":
            stats["blank"] += 1
            continue
        stats["total_data_rows"] += 1
        if kind == "error":
            stats["errors"][temp] = stats["errors"].get(temp, 0) + 1
            continue

        stats["valid"] += 1
        rows.append(envelope(to_epoch_ms_utc(data), float(temp), float(umid), sheet_name, alarm))

        # Validation leg: recompute VPD (refuses implausible temp/RH — an
        # implausible reading dies here, it never reaches the NDJSON) and
        # diff against whatever VPD the sheet itself recorded.
        recomputed = vpd_air_kpa(float(temp), float(umid))
        sheet_vpd = decode_sheet_vpd(vpd_cell)
        if sheet_vpd is None:
            continue
        stats["with_sheet_vpd"] += 1
        diff = recomputed - sheet_vpd
        stats["hist"][sum(1 for e in HIST_EDGES if diff > e)] += 1
        if abs(diff) <= VPD_TOLERANCE_KPA + 1e-9:
            stats["vpd_match"] += 1
        else:
            stats["beyond_tolerance"].append(
                {"row": i, "ts": data.isoformat(), "temp_c": temp, "rh_pct": umid,
                 "sheet_vpd": sheet_vpd, "recomputed": round(recomputed, 4), "diff": round(diff, 4)}
            )
    return rows, stats


HIST_LABELS = ("< -0.02", "-0.02..-0.01", "-0.01..-0.005", "-0.005..+0.005", "+0.005..+0.01", "+0.01..+0.02", "> +0.02")


def print_report(all_stats: list[dict]) -> None:
    for s in all_stats:
        pct = 100.0 * s["vpd_match"] / s["with_sheet_vpd"] if s["with_sheet_vpd"] else float("nan")
        print(f"\n[{s['sheet']}] data rows={s['total_data_rows']} valid={s['valid']} "
              f"skipped-error={sum(s['errors'].values())} blank={s['blank']}")
        if s["errors"]:
            print(f"  error texts: {s['errors']}")
        print(f"  sheet-VPD present: {s['with_sheet_vpd']}  match +/-{VPD_TOLERANCE_KPA} kPa: "
              f"{s['vpd_match']} ({pct:.2f}%)")
        print("  diff histogram (recomputed - sheet, kPa):")
        for label, n in zip(HIST_LABELS, s["hist"]):
            if n:
                print(f"    {label:>16}: {n}")
        for bad in s["beyond_tolerance"]:
            print(f"  BEYOND TOLERANCE: {bad}")


def main(xlsx_path: Path, out_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if wb.epoch != EXCEL_EPOCH_1900:
        raise ValueError(f"workbook epoch {wb.epoch} is not 1899-12-30 — serial conversion assumption broken")
    missing = set(SHEETS) - set(wb.sheetnames)
    if missing:
        raise ValueError(f"expected sheets missing from {xlsx_path}: {sorted(missing)}")

    all_rows: list[dict] = []
    all_stats: list[dict] = []
    seen_ts: set[int] = set()
    for name in SHEETS:
        rows, stats = parse_sheet(wb[name], name)
        dupes = [r["ts"] for r in rows if r["ts"] in seen_ts]
        if dupes:
            # UNIQUE(source_token, ts) in the archive — a duplicate here would
            # be silently collapsed at load time, so it is refused at build time.
            raise ValueError(f"{name}: {len(dupes)} timestamps already emitted by an earlier sheet, e.g. {dupes[:3]}")
        seen_ts.update(r["ts"] for r in rows)
        all_rows.extend(rows)
        all_stats.append(stats)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as f:
        for row in all_rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")

    print_report(all_stats)
    total_vpd = sum(s["with_sheet_vpd"] for s in all_stats)
    total_match = sum(s["vpd_match"] for s in all_stats)
    print(f"\nTOTAL: {len(all_rows)} envelopes -> {out_path}")
    print(f"VPD validation overall: {total_match}/{total_vpd} "
          f"({100.0 * total_match / total_vpd:.2f}%) within +/-{VPD_TOLERANCE_KPA} kPa")
    print("DB load: DEFERRED (archive not live — Felipe gate); this NDJSON is the load-ready artifact.")


DEFAULT_XLSX = Path.home() / "Downloads" / "Thales.xlsx"
DEFAULT_OUT = Path(__file__).parent.parent / "out" / "backfill_thales.ndjson"

if __name__ == "__main__":
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUT
    main(xlsx, out)
